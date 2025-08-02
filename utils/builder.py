"""
DHD Builder

Yohanes Nuwara, August 2025
"""

import json, openpyxl, re, unicodedata
from openpyxl.styles import Font
import glob
import os
from PIL import Image

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

def read_json(json_path):
  with open(json_path, 'r') as f:
      data = json.load(f)
  return data

def slugify(text):
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode()
    text = re.sub(r"[^\w\s-]", "", text).strip().lower()
    return re.sub(r"[-\s]+", "_", text)

def create_dhd_excel(json_path, template_path):
  # Read output json
  data = read_json(json_path)

  # Create safe filename for example well name is written: Acracia # 1
  output_path = f"/content/DHD_{slugify(data['well_name'])}.xlsx"

  # Open the template
  wb = openpyxl.load_workbook(template_path)
  ws = wb.active        # first sheet is the schematic

  # Fill well name in template
  ws["D2"] = data.get("well_name", "")

  units_map = {
      "length_unit":         "E5",
      "depth_unit":          "F5",
      "inner_diameter_unit": "G5",
      "outer_diameter_unit": "H5",
  }
  for key, cell in units_map.items():
      ws[cell] = data.get(key, "") or ""

  # Component list
  start_row = 6
  for i, comp in enumerate(data.get("components", [])):
      r = start_row + i
      ws.cell(r, 3, comp.get("item_id"))
      ws.cell(r, 4, comp.get("name"))
      ws.cell(r, 5, comp.get("length"))
      ws.cell(r, 6, comp.get("depth"))
      ws.cell(r, 7, comp.get("inner_diameter"))
      ws.cell(r, 8, comp.get("outer_diameter"))

  # End of tailpipe
  end_row = start_row + len(data["components"])
  ws.cell(end_row + 1, 4, "END OF TAILPIPE")
  ws.cell(end_row + 1, 6, data["end_of_tailpipe_depth"])


  # End-of-tailpipe depth
  last_depth_row = start_row + len(data["components"])
  ws.cell(last_depth_row + 1, 6, data.get("end_of_tailpipe_depth"))

  wb.save(output_path)
  print("Created:", output_path)

  return output_path

def make_component_dict(json_path, keep_none=True):
    """
    Create component dictionary from extracted DHD table

    Return:
      dict: {"item_id": "name"}
    """
    data = read_json(json_path)

    pairs = (
        (row["item_id"], row["mapped_name"])
        for row in data["components"]
        if keep_none or row["item_id"] is not None
    )

    # Build dict in original order (Python ≥3.7 keeps insertion order)
    return {item_id: mapped for item_id, mapped in pairs}

def official_component_dict(DHD_folder):
  """
  Create official component dictionary
  
  Arguments:
    DHD_folder (str): Path to folder containing individual component images in PNG format

  Returns:
    dict: {"name": "path_to_component_image"}
  """
  image_map = {
      os.path.splitext(os.path.basename(fp))[0]: fp
      for fp in glob.glob(os.path.join(DHD_folder, '*'))
  }  
  return image_map

def stack_dhd(order, mapping, dhd_image_output_path='stack.png', gap_pixels=6, skip_missing=True):
    """
    Create stacked image of DHD components

    Arguments:
    
    order (list): List of DHD components name 
    mapping (dict): Dictionary of DHD component name and image paths
    dhd_image_output_path (str): Path to output image
    gap_pixels (int): Spacing between components when None. Default is 6
    skip_missing (bool): Skip missing component. Default is True

    Returns:
    PIL.Image.Image: Stacked image of DHD components
    """
    items, max_w, total_h = [], 1, 0

    # Pass 1 – gather images & count total height
    for label in order:
        if label is None:                     # spacer
            items.append(('gap', gap_pixels))
            total_h += gap_pixels
            continue

        if label not in mapping:
            if skip_missing:
                continue
            raise KeyError(label)

        src = mapping[label]
        img = Image.open(src) if isinstance(src, str) else src.copy()
        img = img.convert("RGBA")
        items.append(('img', img))
        max_w = max(max_w, img.width)
        total_h += img.height

    # Canvas
    canvas = Image.new("RGBA", (max_w, total_h), (255, 255, 255, 0))

    # Pass 2 – paste images with gaps
    y = 0
    for kind, obj in items:
        if kind == 'gap':
            y += obj                       # just move the cursor
        else:
            img = obj
            x = {'left': 0,
                 'center': (max_w - img.width) // 2,
                 'right': max_w - img.width}['center']
            canvas.alpha_composite(img, (x, y))
            y += img.height

    # Save the image
    canvas.save(dhd_image_output_path)

def apply_dhd_on_excel(excel_in, excel_out, png_path):
  """
  Apply the generated DHD image on Excel

  Arguments:

    excel_in (str): Path to input Excel
    excel_out (str): Path to output Excel
    png_path (str): Path to PNG image
  """
  # open workbook / sheet
  wb = load_workbook(excel_in)
  ws = wb.active 

  # attach the PNG at cell A6
  img = XLImage(png_path)
  img.width  = img.width  / 3
  img.height = img.height / 3
  img.anchor = "A6"  
  ws.add_image(img)

  # save
  wb.save(excel_out)
  print("Created", excel_out) 

def convert_DHD_report_to_excel(pdf_path,
                                dhd_image_path="/content/stack.png",
                                dhd_folder = "/content/drive/MyDrive/GEODATAX ENTERPRISE/TECH & PRODUCT/DHD/DHD IMAGES/DEMO",
                                template_path="/content/drive/MyDrive/GEODATAX ENTERPRISE/TECH & PRODUCT/DHD/DHD Template.xlsx", 
                                api_key="AIzaSyCRZ2-Y86_f9wDZuV_uC6xqvy3Xe3BnQNI"):

  ### 1. Extract JSON from report using Gemini 
  ### During extraction, the original component names will be mapped to official names

  # Define
  prompt_path = '/content/PROMPT.txt'
  output_ocr_path = '/content/JSON'
  model = 'gemini-2.0-flash'

  # Read prompt
  prompt = read_prompt(prompt_path)

  # Run OCR
  run_OCR(pdf_path, prompt, output_ocr_path, model, api_key, display=True)


  ### 2. Convert JSON to Excel, Fill Excel DHD template

  # Define input
  json_path = "/content/JSON/page_1.json"
  excel_output_path = create_dhd_excel(json_path, template_path)  


  ### 3. Extract mapped component name list from JSON

  d = make_component_dict(json_path)
  component_output = list(d.values())


  ### 4. Stack DHD images based on the list

  image_map = official_component_dict(dhd_folder)
  stack_dhd(component_output, image_map, dhd_image_path, gap_pixels=8)


  ### 5. Apply DHD image on Excel
  
  excel_in  = excel_output_path
  excel_out = excel_in
  apply_dhd_on_excel(excel_in, excel_out, dhd_image_path)   
